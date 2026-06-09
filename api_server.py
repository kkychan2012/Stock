"""
Stock Dashboard API Server

Run:  python api_server.py
      python api_server.py --port 5001 --host 0.0.0.0

Endpoints
---------
GET  /                              Dashboard UI

-- Market data (read-only) --
GET  /api/holdings                  Holdings + live P/L from latest price
GET  /api/signals/ma200?days=30     price_gt_ma200 signals (last N days)
GET  /api/signals/ma1030?days=30    ma10_gt_ma30 signals  (last N days)
GET  /api/sold                      Sold positions
GET  /api/monitor                   Monitor list + latest price
GET  /api/prices                    Latest close for all tracked tickers

-- Extraction ticker list --
GET    /api/extraction/tickers      List all tickers
POST   /api/extraction/tickers      Add one ticker  {ticker, notes?}
DELETE /api/extraction/tickers/<t>  Remove one ticker
DELETE /api/extraction/tickers      Clear all tickers
POST   /api/extraction/upload       Bulk upload (CSV/XLSX file or JSON list)
GET    /api/extraction/download     Download ticker list as CSV

-- Fetch trigger --
POST /api/fetch                     Start background fetch  {period?}
GET  /api/fetch/status              Poll fetch progress

Common query params:
  ?ticker=AAPL     Filter by ticker (market data endpoints)
  ?days=N          Signals: how many calendar days back
  ?latest=1        Signals: one most-recent row per ticker
"""

import argparse
import base64
import csv
import io
import os
import subprocess
import sys
import threading
from datetime import datetime, timedelta

from flask import Flask, jsonify, request, render_template, Response, send_file

from db_setup import get_connection, setup_database
from fetch_data import fetch_all, get_tickers_from_db
from scan_patterns import scan_all_patterns, scan_date_range, get_scan_results, get_scan_results_range, get_available_scan_dates
from backtest_engine import run_trading_simulation, calculate_metrics

# ---------------------------------------------------------------------------
# Date normalisation
# ---------------------------------------------------------------------------

_DATE_FORMATS = [
    "%Y-%m-%d",   # 2024-01-31  canonical
    "%d/%m/%Y",   # 31/01/2024  AU/UK
    "%d/%m/%y",   # 31/01/24
    "%m/%d/%Y",   # 01/31/2024  US
    "%m/%d/%y",   # 01/31/24
    "%d-%m-%Y",   # 31-01-2024
    "%d-%m-%y",   # 31-01-24
    "%Y/%m/%d",   # 2024/01/31
    "%d.%m.%Y",   # 31.01.2024
    "%d.%m.%y",   # 31.01.24
    "%d %b %Y",   # 31 Jan 2024
    "%d %B %Y",   # 31 January 2024
    "%b %d, %Y",  # Jan 31, 2024
    "%B %d, %Y",  # January 31, 2024
]

_DATE_COLS = {"buy_date", "sell_date", "signal_date"}


def _normalise_date(s):
    """Return YYYY-MM-DD for any recognised date string, or None."""
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Basic Auth  (enabled only when --user / DASH_USER is set)
# ---------------------------------------------------------------------------

_AUTH_USER: str | None = None
_AUTH_PASS: str        = ""


@app.before_request
def _require_auth():
    if _AUTH_USER is None:
        return                          # auth disabled — local-only mode
    header = request.headers.get("Authorization", "")
    if header.startswith("Basic "):
        try:
            user, _, pw = base64.b64decode(header[6:]).decode().partition(":")
            if user == _AUTH_USER and pw == _AUTH_PASS:
                return                  # correct credentials
        except Exception:
            pass
    return Response(
        "Stock Dashboard — login required.",
        401,
        {"WWW-Authenticate": 'Basic realm="Stock Dashboard"'},
    )


# ---------------------------------------------------------------------------
# Fetch state (module-level, single-user local tool)
# ---------------------------------------------------------------------------

_fetch_state = {"running": False, "total": 0, "done": 0, "log": [], "error": None}
_fetch_lock  = threading.Lock()

_scan_state = {
    "running": False, "total": 0, "done": 0,
    "ticker": "", "error": None, "scan_date": None,
    "mode": "single", "from_date": None, "to_date": None,
}
_scan_lock = threading.Lock()

_insider_scan_state = {"running": False, "log": [], "error": None}
_insider_scan_lock  = threading.Lock()

_INSIDER_PIPELINE = os.path.join(
    os.path.dirname(__file__), "Insider Trading", "insider_pipeline", "main.py"
)

VALID_PERIODS = {"1mo", "3mo", "6mo", "1y", "2y"}

# ---------------------------------------------------------------------------
# Dashboard UI
# ---------------------------------------------------------------------------

@app.get("/")
def index():
    return render_template("dashboard.html")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_LATEST_PRICE_CTE = """
    WITH lp AS (
        SELECT ticker, MAX(date) AS max_date
        FROM stocks_daily
        GROUP BY ticker
    )
"""


def _rows(conn, sql, params=()):
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def _ok(data):
    return jsonify({"count": len(data), "data": data})


def _ticker_filter(alias="h"):
    t = request.args.get("ticker", "").strip().upper()
    if t:
        return f" AND {alias}.ticker = ?", (t,)
    return "", ()


# ---------------------------------------------------------------------------
# GET /api/holdings
# ---------------------------------------------------------------------------

@app.get("/api/holdings")
def get_holdings():
    ticker_sql, ticker_params = _ticker_filter("h")
    sql = f"""
        {_LATEST_PRICE_CTE}
        SELECT
            h.id, h.ticker, h.stock_name, h.avg_buy_price, h.qty,
            h.buy_date, h.notes, h.added_at,
            sd.close                                                    AS current_price,
            sd.date                                                     AS price_date,
            sd.ma6, sd.ma10, sd.ma30, sd.ma50, sd.ma200,
            sd.high_30d, sd.low_30d, sd.vol_ma10,
            sd.pct_change                                               AS day_pct_change,
            sd.direction,
            ROUND((sd.close - h.avg_buy_price) * h.qty, 2)             AS pl_value,
            CASE
                WHEN h.avg_buy_price > 0
                THEN ROUND((sd.close - h.avg_buy_price) / h.avg_buy_price * 100, 2)
                ELSE NULL
            END                                                         AS pl_pct
        FROM holdings h
        LEFT JOIN lp ON h.ticker = lp.ticker
        LEFT JOIN stocks_daily sd ON sd.ticker = lp.ticker AND sd.date = lp.max_date
        WHERE 1=1 {ticker_sql}
        ORDER BY h.ticker
    """
    with get_connection() as conn:
        return _ok(_rows(conn, sql, ticker_params))


# ---------------------------------------------------------------------------
# GET /api/signals/ma200   GET /api/signals/ma1030
# ---------------------------------------------------------------------------

def _get_signals(signal_type: str):
    """Query signals directly from stocks_daily so every historical date is
    visible regardless of when the fetcher was run.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    ago30 = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    date_from   = request.args.get("from",     ago30)
    date_to     = request.args.get("to",       today)
    latest_only = request.args.get("latest",   "0") in ("1", "true", "yes")
    lookback    = request.args.get("lookback", 0, type=int)
    ticker_sql, ticker_params = _ticker_filter("sig")

    # Per-type WHERE conditions (all with explicit alias)
    if signal_type == "price_gt_ma200":
        sig_where_sig  = "sig.close > sig.ma200  AND sig.ma200  IS NOT NULL"
        sig_where_sub  = "sub.close > sub.ma200  AND sub.ma200  IS NOT NULL"
        indicator_col  = "sig.ma200"
        lookback_cond  = "prev.close > prev.ma200 AND prev.ma200 IS NOT NULL"
    else:
        sig_where_sig  = "sig.ma10 > sig.ma30 AND sig.ma10 IS NOT NULL AND sig.ma30 IS NOT NULL"
        sig_where_sub  = "sub.ma10 > sub.ma30 AND sub.ma10 IS NOT NULL AND sub.ma30 IS NOT NULL"
        indicator_col  = "sig.ma10"
        lookback_cond  = "prev.ma10 > prev.ma30 AND prev.ma10 IS NOT NULL"

    # Latest-per-ticker: find the most recent signal date per ticker in range
    if latest_only:
        dedup_cte = f"""
            , dedup AS (
                SELECT sub.ticker, MAX(sub.date) AS max_sig_date
                FROM stocks_daily sub
                WHERE {sig_where_sub}
                  AND sub.date >= ? AND sub.date <= ?
                GROUP BY sub.ticker
            )
        """
        dedup_join   = "JOIN dedup ON sig.ticker = dedup.ticker AND sig.date = dedup.max_sig_date"
        dedup_params = (date_from, date_to)
    else:
        dedup_cte    = ""
        dedup_join   = ""
        dedup_params = ()

    # Fresh-breakout: condition must NOT have been true in the X *trading days*
    # before the signal date.  We find the boundary date by walking back X rows
    # in stocks_daily for that ticker, so weekends and holidays are excluded.
    if lookback > 0:
        lookback_clause = f"""
            AND NOT EXISTS (
                SELECT 1 FROM stocks_daily prev
                WHERE prev.ticker = sig.ticker
                  AND prev.date >= (
                      SELECT MIN(bd.date) FROM (
                          SELECT date FROM stocks_daily
                          WHERE ticker = sig.ticker
                            AND date   < sig.date
                          ORDER BY date DESC
                          LIMIT ?
                      ) bd
                  )
                  AND prev.date < sig.date
                  AND prev.close IS NOT NULL
                  AND {lookback_cond}
            )
        """
        lookback_params = (lookback,)
    else:
        lookback_clause = ""
        lookback_params = ()

    sql = f"""
        {_LATEST_PRICE_CTE}
        {dedup_cte}
        SELECT
            sig.ticker,
            sig.date                 AS signal_date,
            sig.close                AS signal_close,
            {indicator_col}          AS indicator_value,
            cur.close                AS current_price,
            cur.date                 AS current_price_date,
            cur.ma6, cur.ma10, cur.ma30, cur.ma50, cur.ma200,
            cur.high_30d, cur.low_30d,
            cur.pct_change           AS day_pct_change,
            cur.direction
        FROM stocks_daily sig
        {dedup_join}
        LEFT JOIN lp ON sig.ticker = lp.ticker
        LEFT JOIN stocks_daily cur
               ON cur.ticker = lp.ticker AND cur.date = lp.max_date
        WHERE {sig_where_sig}
          AND sig.date >= ? AND sig.date <= ?
          {lookback_clause}
          {ticker_sql}
        ORDER BY sig.date DESC, sig.ticker
    """
    params = dedup_params + (date_from, date_to) + lookback_params + ticker_params
    with get_connection() as conn:
        return _ok(_rows(conn, sql, params))


@app.get("/api/signals/ma200")
def get_signals_ma200():
    return _get_signals("price_gt_ma200")


@app.get("/api/signals/ma1030")
def get_signals_ma1030():
    return _get_signals("ma10_gt_ma30")


# ---------------------------------------------------------------------------
# GET /api/sold
# ---------------------------------------------------------------------------

@app.get("/api/sold")
def get_sold():
    ticker_sql, ticker_params = _ticker_filter("s")
    sql = f"""
        SELECT
            s.id, s.ticker, s.stock_name,
            s.avg_buy_price, s.qty, s.buy_date,
            s.sell_price, s.sell_date,
            s.pl_value, s.notes, s.created_at,
            CASE
                WHEN s.avg_buy_price > 0
                THEN ROUND((s.sell_price - s.avg_buy_price) / s.avg_buy_price * 100, 2)
                ELSE NULL
            END AS pl_pct
        FROM sold s
        WHERE 1=1 {ticker_sql}
        ORDER BY s.sell_date DESC, s.ticker
    """
    with get_connection() as conn:
        return _ok(_rows(conn, sql, ticker_params))


# ---------------------------------------------------------------------------
# Holdings CRUD + upload/download
# ---------------------------------------------------------------------------

def _parse_record_body(body, required):
    """Extract and validate fields from a JSON request body."""
    out = {}
    errors = []
    for field in required:
        val = body.get(field)
        if val is None or str(val).strip() == "":
            errors.append(f"{field} is required")
        else:
            out[field] = val
    if errors:
        return None, errors
    # optional fields
    for field in ("stock_name", "buy_date", "notes", "sell_date"):
        if field in body:
            v = body[field] or None
            out[field] = _normalise_date(v) if field in _DATE_COLS else v
    return out, []


@app.post("/api/holdings")
def add_holding():
    body = request.get_json(silent=True) or {}
    body["ticker"] = (body.get("ticker") or "").strip().upper()
    data, errors = _parse_record_body(body, ["ticker", "avg_buy_price", "qty"])
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    with get_connection() as conn:
        try:
            conn.execute("""
                INSERT INTO holdings (ticker, stock_name, avg_buy_price, qty, buy_date, notes)
                VALUES (:ticker, :stock_name, :avg_buy_price, :qty, :buy_date, :notes)
                ON CONFLICT(ticker) DO UPDATE SET
                    stock_name    = excluded.stock_name,
                    avg_buy_price = excluded.avg_buy_price,
                    qty           = excluded.qty,
                    buy_date      = excluded.buy_date,
                    notes         = excluded.notes
            """, {**{"stock_name": None, "buy_date": None, "notes": None}, **data})
        except Exception as e:
            return jsonify({"error": str(e)}), 400
    return jsonify({"saved": data["ticker"]}), 201


@app.put("/api/holdings/<int:rec_id>")
def update_holding(rec_id):
    body = request.get_json(silent=True) or {}
    if "ticker" in body:
        body["ticker"] = body["ticker"].strip().upper()
    data, errors = _parse_record_body(body, ["ticker", "avg_buy_price", "qty"])
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    with get_connection() as conn:
        cur = conn.execute("""
            UPDATE holdings
            SET ticker        = :ticker,
                stock_name    = :stock_name,
                avg_buy_price = :avg_buy_price,
                qty           = :qty,
                buy_date      = :buy_date,
                notes         = :notes
            WHERE id = :id
        """, {**{"stock_name": None, "buy_date": None, "notes": None}, **data, "id": rec_id})
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"updated": rec_id})


@app.post("/api/holdings/<int:holding_id>/sell")
def sell_holding(holding_id):
    body = request.get_json(silent=True) or {}
    try:
        sell_qty   = float(body.get("sell_qty",   0))
        sell_price = float(body.get("sell_price", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "sell_qty and sell_price must be numbers"}), 400

    sell_date = _normalise_date((body.get("sell_date") or "").strip() or None)
    notes     = (body.get("notes") or "").strip() or None

    if sell_qty <= 0:
        return jsonify({"error": "sell_qty must be greater than 0"}), 400
    if sell_price <= 0:
        return jsonify({"error": "sell_price must be greater than 0"}), 400

    with get_connection() as conn:
        row = conn.execute("SELECT * FROM holdings WHERE id = ?", (holding_id,)).fetchone()
        if not row:
            return jsonify({"error": "holding not found"}), 404
        holding = dict(row)

        if sell_qty > holding["qty"]:
            return jsonify({"error": f"sell_qty ({sell_qty}) exceeds holding qty ({holding['qty']})"}), 400

        pl_value = round((sell_price - (holding["avg_buy_price"] or 0)) * sell_qty, 2)

        conn.execute("""
            INSERT INTO sold (ticker, stock_name, avg_buy_price, qty, buy_date,
                              sell_price, sell_date, pl_value, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (holding["ticker"], holding["stock_name"], holding["avg_buy_price"],
              sell_qty, holding["buy_date"], sell_price, sell_date, pl_value, notes))

        remaining = holding["qty"] - sell_qty
        if remaining <= 0:
            conn.execute("DELETE FROM holdings WHERE id = ?", (holding_id,))
        else:
            conn.execute("UPDATE holdings SET qty = ? WHERE id = ?", (remaining, holding_id))

    return jsonify({"sold": holding["ticker"], "remaining_qty": max(0, remaining)}), 201


@app.delete("/api/holdings/<int:rec_id>")
def delete_holding(rec_id):
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM holdings WHERE id = ?", (rec_id,))
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"deleted": rec_id})


@app.post("/api/holdings/upload")
def upload_holdings():
    return _upload_table(
        request,
        table="holdings",
        columns=["ticker", "stock_name", "avg_buy_price", "qty", "buy_date", "notes"],
        upsert=True,
    )


@app.get("/api/holdings/download")
def download_holdings():
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT ticker, stock_name, avg_buy_price, qty, buy_date, notes FROM holdings ORDER BY ticker"
        ).fetchall()
    return _csv_response(rows, "holdings.csv")


# ---------------------------------------------------------------------------
# Sold CRUD + upload/download
# ---------------------------------------------------------------------------

@app.post("/api/sold")
def add_sold():
    body = request.get_json(silent=True) or {}
    body["ticker"] = (body.get("ticker") or "").strip().upper()
    data, errors = _parse_record_body(
        body, ["ticker", "avg_buy_price", "qty", "sell_price", "sell_date"]
    )
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    # Auto-calculate pl_value if not provided
    if "pl_value" not in data or data.get("pl_value") is None:
        try:
            data["pl_value"] = round(
                (float(data["sell_price"]) - float(data["avg_buy_price"])) * float(data["qty"]), 2
            )
        except (TypeError, ValueError):
            data["pl_value"] = None
    with get_connection() as conn:
        cur = conn.execute("""
            INSERT INTO sold (ticker, stock_name, avg_buy_price, qty, buy_date,
                              sell_price, sell_date, pl_value, notes)
            VALUES (:ticker, :stock_name, :avg_buy_price, :qty, :buy_date,
                    :sell_price, :sell_date, :pl_value, :notes)
        """, {**{"stock_name": None, "buy_date": None, "notes": None, "pl_value": None}, **data})
        new_id = cur.lastrowid
    return jsonify({"saved": new_id}), 201


@app.put("/api/sold/<int:rec_id>")
def update_sold(rec_id):
    body = request.get_json(silent=True) or {}
    if "ticker" in body:
        body["ticker"] = body["ticker"].strip().upper()
    data, errors = _parse_record_body(
        body, ["ticker", "avg_buy_price", "qty", "sell_price", "sell_date"]
    )
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    try:
        data["pl_value"] = round(
            (float(data["sell_price"]) - float(data["avg_buy_price"])) * float(data["qty"]), 2
        )
    except (TypeError, ValueError):
        data["pl_value"] = None
    with get_connection() as conn:
        cur = conn.execute("""
            UPDATE sold
            SET ticker        = :ticker,
                stock_name    = :stock_name,
                avg_buy_price = :avg_buy_price,
                qty           = :qty,
                buy_date      = :buy_date,
                sell_price    = :sell_price,
                sell_date     = :sell_date,
                pl_value      = :pl_value,
                notes         = :notes
            WHERE id = :id
        """, {**{"stock_name": None, "buy_date": None, "notes": None}, **data, "id": rec_id})
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"updated": rec_id})


@app.delete("/api/sold/<int:rec_id>")
def delete_sold(rec_id):
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM sold WHERE id = ?", (rec_id,))
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"deleted": rec_id})


@app.post("/api/sold/upload")
def upload_sold():
    return _upload_table(
        request,
        table="sold",
        columns=["ticker", "stock_name", "avg_buy_price", "qty", "buy_date",
                 "sell_price", "sell_date", "pl_value", "notes"],
        upsert=False,
    )


@app.get("/api/sold/download")
def download_sold():
    with get_connection() as conn:
        rows = conn.execute("""
            SELECT ticker, stock_name, avg_buy_price, qty, buy_date,
                   sell_price, sell_date, pl_value, notes
            FROM sold ORDER BY sell_date DESC, ticker
        """).fetchall()
    return _csv_response(rows, "sold.csv")


# ---------------------------------------------------------------------------
# Shared upload / download helpers
# ---------------------------------------------------------------------------

def _read_upload_rows(req):
    """Parse CSV or XLSX from a multipart file upload. Returns list of dicts."""
    if "file" not in req.files:
        return None, "No file in request"
    f    = req.files["file"]
    name = (f.filename or "").lower()
    if name.endswith(".csv"):
        text   = f.read().decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader], None
    elif name.endswith((".xlsx", ".xls")):
        import openpyxl
        wb  = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws  = wb.active
        headers = None
        rows = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip().lower() if c else "" for c in row]
            else:
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        wb.close()
        return rows, None
    else:
        return None, "Only .csv and .xlsx files are supported"


def _upload_table(req, table, columns, upsert):
    rows, err = _read_upload_rows(req)
    if err:
        return jsonify({"error": err}), 400
    if not rows:
        return jsonify({"error": "No data rows found in file"}), 400

    added = skipped = 0
    with get_connection() as conn:
        for row in rows:
            # Normalise keys to lowercase
            row = {k.lower().strip(): v for k, v in row.items()}
            ticker = (row.get("ticker") or "").strip().upper()
            if not ticker:
                skipped += 1
                continue

            vals = {"ticker": ticker}
            for col in columns:
                if col == "ticker":
                    continue
                v = row.get(col, "")
                vals[col] = v.strip() if isinstance(v, str) else v or None
                if vals[col] == "":
                    vals[col] = None
                if col in _DATE_COLS and vals[col]:
                    vals[col] = _normalise_date(vals[col])

            # Auto-calc pl_value for sold if missing
            if table == "sold" and not vals.get("pl_value"):
                try:
                    vals["pl_value"] = round(
                        (float(vals["sell_price"]) - float(vals["avg_buy_price"])) * float(vals["qty"]), 2
                    )
                except (TypeError, ValueError):
                    vals["pl_value"] = None

            col_list = ", ".join(columns)
            placeholders = ", ".join(f":{c}" for c in columns)

            if upsert:
                update_set = ", ".join(
                    f"{c} = excluded.{c}" for c in columns if c != "ticker"
                )
                sql = f"""
                    INSERT INTO {table} ({col_list}) VALUES ({placeholders})
                    ON CONFLICT(ticker) DO UPDATE SET {update_set}
                """
            else:
                sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"

            try:
                conn.execute(sql, vals)
                added += 1
            except Exception:
                skipped += 1

    return jsonify({"added": added, "skipped": skipped})


def _csv_response(rows, filename):
    if not rows:
        buf = io.StringIO()
        return Response(buf.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(rows[0].keys())
    for r in rows:
        writer.writerow([r[k] if r[k] is not None else "" for k in r.keys()])
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------------------------------------------------------------------------
# GET /api/monitor
# ---------------------------------------------------------------------------

@app.get("/api/monitor")
def get_monitor():
    ticker_sql, ticker_params = _ticker_filter("m")
    sql = f"""
        {_LATEST_PRICE_CTE}
        SELECT
            m.id, m.ticker, m.stock_name, m.reason, m.comment,
            m.signal_date, m.signal_price, m.added_at,
            sd.close         AS current_price,
            sd.date          AS price_date,
            sd.ma6, sd.ma10, sd.ma30, sd.ma50, sd.ma200,
            sd.high_30d, sd.low_30d, sd.vol_ma10,
            sd.pct_change    AS day_pct_change,
            sd.direction,
            CASE
                WHEN m.signal_price > 0 AND sd.close IS NOT NULL
                THEN ROUND((sd.close - m.signal_price) / m.signal_price * 100, 2)
                ELSE NULL
            END AS since_signal_pct
        FROM monitor_list m
        LEFT JOIN lp ON m.ticker = lp.ticker
        LEFT JOIN stocks_daily sd ON sd.ticker = lp.ticker AND sd.date = lp.max_date
        WHERE 1=1 {ticker_sql}
        ORDER BY m.ticker
    """
    with get_connection() as conn:
        return _ok(_rows(conn, sql, ticker_params))


# ---------------------------------------------------------------------------
# Monitor list CRUD + upload/download
# ---------------------------------------------------------------------------

@app.get("/api/monitor/tickers")
def list_monitor_tickers():
    """Lightweight endpoint — returns just the ticker strings in the monitor list."""
    with get_connection() as conn:
        rows = conn.execute("SELECT ticker FROM monitor_list ORDER BY ticker").fetchall()
    return jsonify({"tickers": [r["ticker"] for r in rows]})


@app.post("/api/monitor")
def add_monitor():
    body         = request.get_json(silent=True) or {}
    ticker       = (body.get("ticker")       or "").strip().upper()
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    stock_name   = (body.get("stock_name")   or "").strip() or None
    reason       = (body.get("reason")       or "").strip() or None
    comment      = (body.get("comment")      or "").strip() or None
    signal_date  = _normalise_date((body.get("signal_date") or "").strip() or None)
    signal_price = body.get("signal_price")
    try:
        signal_price = float(signal_price) if signal_price is not None else None
    except (TypeError, ValueError):
        signal_price = None
    with get_connection() as conn:
        try:
            conn.execute(
                """INSERT INTO monitor_list (ticker, stock_name, reason, comment, signal_date, signal_price)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (ticker, stock_name, reason, comment, signal_date, signal_price)
            )
        except Exception:
            return jsonify({"error": f"{ticker} already in monitor list"}), 409
    return jsonify({"added": ticker}), 201


@app.put("/api/monitor/<int:rec_id>")
def update_monitor(rec_id):
    body         = request.get_json(silent=True) or {}
    ticker       = (body.get("ticker")       or "").strip().upper()
    stock_name   = (body.get("stock_name")   or "").strip() or None
    reason       = (body.get("reason")       or "").strip() or None
    comment      = (body.get("comment")      or "").strip() or None
    signal_date  = _normalise_date((body.get("signal_date") or "").strip() or None)
    signal_price = body.get("signal_price")
    try:
        signal_price = float(signal_price) if signal_price is not None and str(signal_price).strip() != "" else None
    except (TypeError, ValueError):
        signal_price = None
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    with get_connection() as conn:
        cur = conn.execute(
            """UPDATE monitor_list
               SET ticker=?, stock_name=?, reason=?, comment=?, signal_date=?, signal_price=?
               WHERE id=?""",
            (ticker, stock_name, reason, comment, signal_date, signal_price, rec_id)
        )
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"updated": rec_id})


@app.delete("/api/monitor/<int:rec_id>")
def delete_monitor(rec_id):
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM monitor_list WHERE id = ?", (rec_id,))
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"deleted": rec_id})


@app.post("/api/monitor/upload")
def upload_monitor():
    return _upload_table(
        request,
        table="monitor_list",
        columns=["ticker", "stock_name", "reason", "comment"],
        upsert=True,
    )


@app.get("/api/monitor/download")
def download_monitor():
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT ticker, stock_name, reason, comment, signal_date, signal_price, added_at
               FROM monitor_list ORDER BY ticker"""
        ).fetchall()
    return _csv_response(rows, "monitor_list.csv")


# ---------------------------------------------------------------------------
# GET /api/insider   — insider signals stored by the pipeline
# ---------------------------------------------------------------------------

@app.get("/api/insider")
def get_insider():
    ticker_sql, ticker_params = _ticker_filter("s")
    from_date = (request.args.get("from") or "").strip()
    to_date   = (request.args.get("to")   or "").strip()
    date_sql, date_params = "", ()
    if from_date:
        date_sql  += " AND s.transaction_date >= ?"
        date_params += (from_date,)
    if to_date:
        date_sql  += " AND s.transaction_date <= ?"
        date_params += (to_date,)
    sql = f"""
        SELECT
            s.id, s.filed_date, s.transaction_date,
            s.ticker, s.company_name, s.insider_name, s.role,
            s.transaction_type, s.shares, s.price, s.total_value,
            s.cluster_buy, s.flag_10b51, s.filing_url, s.source,
            s.scan_run_at
        FROM insider_signals s
        WHERE 1=1 {ticker_sql} {date_sql}
        ORDER BY s.transaction_date DESC, s.total_value DESC
    """
    with get_connection() as conn:
        return _ok(_rows(conn, sql, ticker_params + date_params))


@app.delete("/api/insider/<int:rec_id>")
def delete_insider(rec_id):
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM insider_signals WHERE id = ?", (rec_id,))
    if cur.rowcount == 0:
        return jsonify({"error": "record not found"}), 404
    return jsonify({"deleted": rec_id})


# ---------------------------------------------------------------------------
# POST /api/insider/scan   GET /api/insider/scan/status
# ---------------------------------------------------------------------------

@app.post("/api/insider/scan")
def start_insider_scan():
    with _insider_scan_lock:
        if _insider_scan_state["running"]:
            return jsonify({"status": "already_running"}), 409

        body      = request.get_json(silent=True) or {}
        tickers   = (body.get("tickers") or "").strip()
        from_date = (body.get("from_date") or "").strip()
        to_date   = (body.get("to_date")   or datetime.now().strftime("%Y-%m-%d")).strip()

        if not tickers:
            return jsonify({"error": "tickers is required (e.g. NOK or NOK,ERIC)"}), 400
        if not from_date:
            return jsonify({"error": "from_date is required (YYYY-MM-DD)"}), 400

        _insider_scan_state.update({"running": True, "log": [], "error": None})

    cmd = [sys.executable, _INSIDER_PIPELINE,
           "--ticker", tickers, "--from", from_date, "--to", to_date]

    def _run():
        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                cwd=os.path.dirname(_INSIDER_PIPELINE),
            )
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    _insider_scan_state["log"].append(line)
            proc.wait()
            if proc.returncode != 0:
                _insider_scan_state["error"] = f"Pipeline exited with code {proc.returncode}"
        except Exception as exc:
            _insider_scan_state["error"] = str(exc)
        finally:
            _insider_scan_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started"})


@app.get("/api/insider/scan/status")
def insider_scan_status():
    return jsonify({
        "running": _insider_scan_state["running"],
        "log":     list(_insider_scan_state["log"]),
        "error":   _insider_scan_state["error"],
    })


# ---------------------------------------------------------------------------
# GET /api/prices
# ---------------------------------------------------------------------------

@app.get("/api/prices")
def get_prices():
    ticker_sql, ticker_params = _ticker_filter("t")
    sql = f"""
        {_LATEST_PRICE_CTE}
        , all_tickers AS (
            SELECT ticker FROM holdings
            UNION
            SELECT ticker FROM monitor_list
        )
        SELECT
            t.ticker,
            sd.close         AS price,
            sd.date          AS price_date,
            sd.open, sd.high, sd.low,
            sd.volume, sd.vol_ma10,
            sd.pct_change    AS day_pct_change,
            sd.direction
        FROM all_tickers t
        LEFT JOIN lp ON t.ticker = lp.ticker
        LEFT JOIN stocks_daily sd ON sd.ticker = lp.ticker AND sd.date = lp.max_date
        WHERE 1=1 {ticker_sql}
        ORDER BY t.ticker
    """
    with get_connection() as conn:
        return _ok(_rows(conn, sql, ticker_params))


# ---------------------------------------------------------------------------
# Extraction ticker list  CRUD
# ---------------------------------------------------------------------------

@app.get("/api/extraction/tickers")
def list_extraction_tickers():
    with get_connection() as conn:
        rows = _rows(conn, "SELECT id, ticker, notes, added_at FROM extraction_tickers ORDER BY ticker")
    return jsonify({"count": len(rows), "tickers": rows})


@app.post("/api/extraction/tickers")
def add_extraction_ticker():
    body   = request.get_json(silent=True) or {}
    ticker = (body.get("ticker") or "").strip().upper()
    notes  = (body.get("notes") or "").strip() or None
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    with get_connection() as conn:
        try:
            conn.execute(
                "INSERT INTO extraction_tickers (ticker, notes) VALUES (?, ?)",
                (ticker, notes)
            )
        except Exception:
            return jsonify({"error": f"{ticker} already exists"}), 409
    return jsonify({"added": ticker}), 201


@app.delete("/api/extraction/tickers/<ticker>")
def delete_extraction_ticker(ticker):
    ticker = ticker.strip().upper()
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM extraction_tickers WHERE ticker = ?", (ticker,))
    if cur.rowcount == 0:
        return jsonify({"error": f"{ticker} not found"}), 404
    return jsonify({"deleted": ticker})


@app.delete("/api/extraction/tickers")
def clear_extraction_tickers():
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM extraction_tickers")
    return jsonify({"deleted": cur.rowcount})


# ---------------------------------------------------------------------------
# POST /api/extraction/upload  — bulk CSV / XLSX / JSON
# ---------------------------------------------------------------------------

@app.post("/api/extraction/upload")
def upload_extraction_tickers():
    tickers_to_add = []

    # --- JSON body path ---
    if request.is_json:
        body = request.get_json(silent=True) or {}
        raw  = body.get("tickers", [])
        tickers_to_add = [t.strip().upper() for t in raw if str(t).strip()]

    # --- File upload path ---
    elif "file" in request.files:
        f    = request.files["file"]
        name = (f.filename or "").lower()

        if name.endswith(".csv"):
            text    = f.read().decode("utf-8", errors="replace")
            reader  = csv.reader(io.StringIO(text))
            for row in reader:
                if row:
                    val = row[0].strip().upper()
                    if val and val != "TICKER":
                        tickers_to_add.append(val)

        elif name.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                val = str(row[0] or "").strip().upper()
                if val and val != "TICKER":
                    tickers_to_add.append(val)
            wb.close()
        else:
            return jsonify({"error": "Only .csv and .xlsx files are supported"}), 400
    else:
        return jsonify({"error": "Send a file (CSV/XLSX) or JSON {\"tickers\":[...]}"}), 400

    if not tickers_to_add:
        return jsonify({"error": "No valid tickers found in upload"}), 400

    added = skipped = 0
    with get_connection() as conn:
        for t in tickers_to_add:
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO extraction_tickers (ticker) VALUES (?)", (t,)
                )
                if conn.execute(
                    "SELECT changes()"
                ).fetchone()[0]:
                    added += 1
                else:
                    skipped += 1
            except Exception:
                skipped += 1

    return jsonify({"added": added, "skipped": skipped})


# ---------------------------------------------------------------------------
# GET /api/extraction/download  — download ticker list as CSV
# ---------------------------------------------------------------------------

@app.get("/api/extraction/download")
def download_extraction_tickers():
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT ticker, notes, added_at FROM extraction_tickers ORDER BY ticker"
        ).fetchall()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["ticker", "notes", "added_at"])
    for r in rows:
        writer.writerow([r["ticker"], r["notes"] or "", r["added_at"]])

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=tickers.csv"},
    )


# ---------------------------------------------------------------------------
# POST /api/fetch   GET /api/fetch/status
# ---------------------------------------------------------------------------

@app.post("/api/fetch")
def start_fetch():
    with _fetch_lock:
        if _fetch_state["running"]:
            return jsonify({"status": "already_running"}), 409

        body   = request.get_json(silent=True) or {}
        period = body.get("period", "2y")
        if period not in VALID_PERIODS:
            return jsonify({"error": f"Invalid period. Use one of: {sorted(VALID_PERIODS)}"}), 400

        tickers = get_tickers_from_db()
        if not tickers:
            return jsonify({"error": "extraction_tickers table is empty. Add tickers first."}), 400

        _fetch_state.update({
            "running": True,
            "total":   len(tickers),
            "done":    0,
            "log":     [],
            "error":   None,
        })

    def _run():
        def _cb(msg: str):
            _fetch_state["log"].append(msg)
            if msg.startswith(("OK ", "SKIP ", "ERROR ")):
                _fetch_state["done"] += 1

        try:
            fetch_all(tickers, period=period, progress_cb=_cb)
        except Exception as exc:
            _fetch_state["error"] = str(exc)
        finally:
            _fetch_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started", "total": len(tickers)})


@app.get("/api/fetch/status")
def fetch_status():
    return jsonify({
        "running": _fetch_state["running"],
        "total":   _fetch_state["total"],
        "done":    _fetch_state["done"],
        "log":     list(_fetch_state["log"]),
        "error":   _fetch_state["error"],
    })


# ---------------------------------------------------------------------------
# Pattern scanner  POST /api/patterns/scan   GET /api/patterns/scan/status
#                  GET  /api/patterns/results  GET /api/patterns/dates
# ---------------------------------------------------------------------------

@app.post("/api/patterns/scan")
def start_pattern_scan():
    with _scan_lock:
        if _scan_state["running"]:
            return jsonify({"status": "already_running"}), 409
        body      = request.get_json(silent=True) or {}
        from_date = (body.get("from_date") or "").strip() or None
        to_date   = (body.get("to_date")   or "").strip() or None
        scan_date = (body.get("date")       or "").strip() or None
        is_range  = bool(from_date and to_date)

        if is_range:
            _scan_state.update({
                "running": True, "total": 0, "done": 0, "ticker": "",
                "error": None, "scan_date": to_date,
                "mode": "range", "from_date": from_date, "to_date": to_date,
            })
        else:
            scan_date = scan_date or datetime.now().strftime("%Y-%m-%d")
            _scan_state.update({
                "running": True, "total": 0, "done": 0, "ticker": "",
                "error": None, "scan_date": scan_date,
                "mode": "single", "from_date": None, "to_date": None,
            })

    def _run():
        def _cb(done, total, ticker):
            _scan_state["done"]   = done
            _scan_state["total"]  = total
            _scan_state["ticker"] = ticker
        try:
            if is_range:
                scan_date_range(from_date, to_date, progress_cb=_cb)
            else:
                scan_all_patterns(scan_date, progress_cb=_cb)
        except Exception as exc:
            _scan_state["error"] = str(exc)
        finally:
            _scan_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    if is_range:
        return jsonify({"status": "started", "from_date": from_date, "to_date": to_date, "mode": "range"})
    return jsonify({"status": "started", "scan_date": scan_date, "mode": "single"})


@app.get("/api/patterns/scan/status")
def pattern_scan_status():
    return jsonify({k: _scan_state[k] for k in
                    ("running", "total", "done", "ticker", "error", "scan_date",
                     "mode", "from_date", "to_date")})


@app.get("/api/patterns/results")
def get_pattern_results_api():
    from_date = (request.args.get("from") or "").strip()
    to_date   = (request.args.get("to")   or "").strip()
    if from_date and to_date:
        results = get_scan_results_range(from_date, to_date)
        return jsonify({"count": len(results), "data": results,
                        "from_date": from_date, "to_date": to_date})
    scan_date = (request.args.get("date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
    results   = get_scan_results(scan_date)
    return jsonify({"count": len(results), "data": results, "scan_date": scan_date})


@app.get("/api/patterns/dates")
def get_pattern_dates_api():
    return jsonify({"dates": get_available_scan_dates()})


# ---------------------------------------------------------------------------
# Backtest  POST /api/backtest/single   POST /api/backtest/batch
# ---------------------------------------------------------------------------

def _parse_rules(body):
    """Extract strategy percentage rules from request body (values as %)."""
    return {
        "t1":    float(body.get("t1",    10.0)) / 100.0,
        "sl":    float(body.get("sl",    11.0)) / 100.0,
        "t2":    float(body.get("t2",    20.0)) / 100.0,
        "rev":   float(body.get("rev",    0.0)) / 100.0,
        "prot":  float(body.get("prot",  10.0)) / 100.0,
        "trail": float(body.get("trail", 10.0)) / 100.0,
    }


@app.post("/api/backtest/single")
def backtest_single():
    body   = request.get_json(silent=True) or {}
    ticker = (body.get("ticker") or "").strip().upper()
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400

    try:
        p0         = float(body["p0"])
        start_date = (body.get("start_date") or "").strip()
        end_date   = (body.get("end_date")   or "").strip() or None
        rules      = _parse_rules(body)
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({"error": f"Invalid parameters: {exc}"}), 400

    if not start_date:
        return jsonify({"error": "start_date is required"}), 400

    with get_connection() as conn:
        if end_date:
            rows = conn.execute(
                "SELECT date, open, high, low, high_30d FROM stocks_daily"
                " WHERE ticker=? AND date>=? AND date<=? ORDER BY date",
                (ticker, start_date, end_date),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT date, open, high, low, high_30d FROM stocks_daily"
                " WHERE ticker=? AND date>=? ORDER BY date",
                (ticker, start_date),
            ).fetchall()

    if not rows:
        return jsonify({"error": f"No price data for {ticker} from {start_date}"}), 404

    txs = run_trading_simulation([dict(r) for r in rows], p0, start_date, rules)
    if not txs:
        return jsonify({"error": "Simulation produced no transactions"}), 404

    initial_cost, total_pnl, roi = calculate_metrics(txs)
    return jsonify({
        "ticker":       ticker,
        "initial_cost": round(initial_cost, 2),
        "total_pnl":    round(total_pnl,    2),
        "roi":          round(roi,           2),
        "status":       txs[-1]["action"],
        "transactions": txs,
    })


@app.post("/api/backtest/selection")
def backtest_selection():
    body  = request.get_json(silent=True) or {}
    items = body.get("items") or []
    if not items:
        return jsonify({"error": "items list is required"}), 400
    try:
        rules = _parse_rules(body)
    except (ValueError, TypeError) as exc:
        return jsonify({"error": f"Invalid parameters: {exc}"}), 400

    results = []
    for item in items:
        ticker     = (item.get("ticker") or "").strip().upper()
        if not ticker:
            continue
        try:
            p0         = float(item["p0"])
            start_date = str(item.get("start_date") or "").strip()
        except (KeyError, ValueError, TypeError):
            results.append({"ticker": ticker, "error": "Invalid p0 or start_date"})
            continue

        if p0 <= 0:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "p0 must be > 0"})
            continue
        if not start_date:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "start_date is required"})
            continue

        with get_connection() as conn:
            rows = conn.execute(
                "SELECT date, open, high, low, high_30d FROM stocks_daily"
                " WHERE ticker=? AND date>=? ORDER BY date",
                (ticker, start_date),
            ).fetchall()

        if not rows:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "No price data found",
                             "initial_cost": None, "total_pnl": None,
                             "roi": None, "status": None, "transactions": []})
            continue

        txs = run_trading_simulation([dict(r) for r in rows], p0, start_date, rules)
        if not txs:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "Simulation produced no transactions",
                             "initial_cost": None, "total_pnl": None,
                             "roi": None, "status": None, "transactions": []})
            continue

        initial_cost, total_pnl, roi = calculate_metrics(txs)
        results.append({
            "ticker":       ticker,
            "p0":           p0,
            "start_date":   start_date,
            "initial_cost": round(initial_cost, 2),
            "total_pnl":    round(total_pnl,    2),
            "roi":          round(roi,           2),
            "status":       txs[-1]["action"],
            "transactions": txs,
        })

    return jsonify({"count": len(results), "data": results})


@app.post("/api/backtest/batch")
def backtest_batch():
    body = request.get_json(silent=True) or {}
    try:
        rules = _parse_rules(body)
    except (ValueError, TypeError) as exc:
        return jsonify({"error": f"Invalid parameters: {exc}"}), 400

    with get_connection() as conn:
        holdings = conn.execute(
            "SELECT ticker, avg_buy_price, buy_date FROM holdings ORDER BY ticker"
        ).fetchall()

    if not holdings:
        return jsonify({"error": "No holdings found — add holdings first"}), 404

    results = []
    for h in holdings:
        ticker     = h["ticker"]
        p0         = float(h["avg_buy_price"] or 0)
        start_date = h["buy_date"] or datetime.now().strftime("%Y-%m-%d")

        if p0 <= 0:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "avg_buy_price is 0 or missing"})
            continue

        with get_connection() as conn:
            rows = conn.execute(
                "SELECT date, open, high, low, high_30d FROM stocks_daily"
                " WHERE ticker=? AND date>=? ORDER BY date",
                (ticker, start_date),
            ).fetchall()

        if not rows:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "No price data found",
                             "initial_cost": None, "total_pnl": None,
                             "roi": None, "status": None, "transactions": []})
            continue

        txs = run_trading_simulation([dict(r) for r in rows], p0, start_date, rules)
        if not txs:
            results.append({"ticker": ticker, "p0": p0, "start_date": start_date,
                             "error": "Simulation produced no transactions",
                             "initial_cost": None, "total_pnl": None,
                             "roi": None, "status": None, "transactions": []})
            continue

        initial_cost, total_pnl, roi = calculate_metrics(txs)
        results.append({
            "ticker":       ticker,
            "p0":           p0,
            "start_date":   start_date,
            "initial_cost": round(initial_cost, 2),
            "total_pnl":    round(total_pnl,    2),
            "roi":          round(roi,           2),
            "status":       txs[-1]["action"],
            "transactions": txs,
        })

    return jsonify({"count": len(results), "data": results})


# ---------------------------------------------------------------------------
# Backup / Restore
# ---------------------------------------------------------------------------

@app.get("/api/backup/download")
def backup_download():
    from db_setup import DB_PATH
    if not os.path.exists(DB_PATH):
        return jsonify({"error": "Database file not found"}), 404
    return send_file(DB_PATH, as_attachment=True, download_name="stock_dashboard.db",
                     mimetype="application/octet-stream")


@app.post("/api/backup/restore")
def backup_restore():
    from db_setup import DB_PATH
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    header = f.read(16)
    if not header.startswith(b"SQLite format 3"):
        return jsonify({"error": "Not a valid SQLite database file"}), 400
    f.seek(0)
    # Write to a temp file first, then atomically replace
    tmp_path = DB_PATH + ".restore_tmp"
    try:
        f.save(tmp_path)
        os.replace(tmp_path, DB_PATH)
    except Exception as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "message": "Database restored. Please refresh the page."})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Stock Dashboard API")
    parser.add_argument("--host",  default="127.0.0.1")
    parser.add_argument("--port",  type=int, default=5000)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--user",  default=os.environ.get("DASH_USER", ""),
                        help="Enable Basic Auth with this username")
    parser.add_argument("--pass",  dest="password",
                        default=os.environ.get("DASH_PASS", ""),
                        help="Basic Auth password")
    args = parser.parse_args()

    if args.user:
        _AUTH_USER = args.user
        _AUTH_PASS = args.password
        print(f"  Basic Auth ENABLED  (user: {_AUTH_USER})")
    else:
        print("  Basic Auth disabled — use --user / --pass to protect the dashboard")

    setup_database()
    print(f"\nStock Dashboard API running at http://{args.host}:{args.port}")
    print("Endpoints:")
    print("  GET  /api/holdings")
    print("  GET  /api/signals/ma200?days=30")
    print("  GET  /api/signals/ma1030?days=30")
    print("  GET  /api/sold  |  /api/monitor  |  /api/prices")
    print("  GET  /api/extraction/tickers  (list)")
    print("  POST /api/extraction/tickers  (add one)")
    print("  POST /api/extraction/upload   (bulk CSV/XLSX/JSON)")
    print("  GET  /api/extraction/download (download CSV)")
    print("  POST /api/fetch               (start fetch)")
    print("  GET  /api/fetch/status        (poll progress)\n")
    app.run(host=args.host, port=args.port, debug=args.debug, threaded=True)
