"""
Build the multi-sheet Excel report with formatting.
"""
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter

from config import OUTPUT_FILE

logger = logging.getLogger(__name__)

# Colour fills
RED_FILL = PatternFill("solid", fgColor="FF4444")
ORANGE_FILL = PatternFill("solid", fgColor="FFA500")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True)

DISPLAY_COLUMNS = [
    "filed_date", "transaction_date", "ticker", "company_name",
    "insider_name", "role", "transaction_type", "shares",
    "price", "total_value", "flag_10b51", "cluster_buy", "source", "filing_url",
]
COLUMN_HEADERS = [
    "Filed Date", "Transaction Date", "Ticker", "Company Name",
    "Insider Name", "Role", "Transaction Type", "Shares",
    "Price (USD)", "Total Value (USD)", "10b5-1 Flag",
    "Cluster Buy Flag", "Source", "SEC Filing URL",
]
DOLLAR_COLS = {"Price (USD)", "Total Value (USD)"}
DATE_COLS = {"Filed Date", "Transaction Date"}


def _write_df_to_sheet(ws, df: pd.DataFrame, large_buy_threshold: float = 1_000_000):
    """Write a DataFrame to a worksheet with formatting."""
    # Header row
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, record in enumerate(df[DISPLAY_COLUMNS].itertuples(index=False), 2):
        row_data = list(record)
        is_cluster = bool(row_data[DISPLAY_COLUMNS.index("cluster_buy")])
        total_val = row_data[DISPLAY_COLUMNS.index("total_value")]
        try:
            total_val = float(total_val)
        except (TypeError, ValueError):
            total_val = 0.0

        if is_cluster:
            fill = RED_FILL
        elif total_val > large_buy_threshold:
            fill = ORANGE_FILL
        else:
            fill = WHITE_FILL

        for col_idx, (header, value) in enumerate(zip(COLUMN_HEADERS, row_data), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

            if header in DATE_COLS and hasattr(value, "date"):
                cell.value = value.date()
                cell.number_format = "YYYY-MM-DD"
            elif header in DOLLAR_COLS:
                try:
                    cell.value = float(value)
                    cell.number_format = '$#,##0'
                except (TypeError, ValueError):
                    cell.value = value
            elif header == "Shares":
                try:
                    cell.value = int(float(value))
                    cell.number_format = "#,##0"
                except (TypeError, ValueError):
                    cell.value = value
            elif header in ("10b5-1 Flag", "Cluster Buy Flag"):
                cell.value = "YES" if value else "NO"
                if value:
                    cell.font = Font(bold=True, color="CC0000")
            else:
                cell.value = str(value) if value is not None else ""

    # Freeze header, auto-filter, auto-fit columns
    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        # Sample widths
        max_len = max(len(header), 12)
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 200),
                                 min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    val_len = len(str(cell.value))
                    max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.row_dimensions[1].height = 20


def _build_dashboard(ws, df: pd.DataFrame, scanned_count: int):
    ws.title = "Dashboard"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cluster_buys = df[df["cluster_buy"]].copy() if "cluster_buy" in df.columns and not df.empty else pd.DataFrame()

    stats = [
        ("Last Updated", now),
        ("Total Filings Scanned", scanned_count),
        ("Qualifying Buys Found", len(df)),
        ("Cluster Buy Events", len(cluster_buys)),
        ("Total Insider Buy Value", f"${df['total_value'].sum():,.0f}"),
        ("Large Buys (>$1M)", int((df["total_value"] > 1_000_000).sum())),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A1"

    row = 1
    header_cell = ws.cell(row=row, column=1, value="Insider Trading Pipeline — Summary")
    header_cell.font = Font(bold=True, size=14)
    header_cell.fill = HEADER_FILL
    header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells("A1:B1")
    row += 1

    for label, value in stats:
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = BOLD_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top 10 Companies by Insider Buy Value").font = BOLD_FONT
    row += 1

    ws.cell(row=row, column=1, value="Ticker").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Total Buy Value (USD)").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1

    if not df.empty:
        top10 = (df.groupby("ticker")["total_value"]
                 .sum()
                 .sort_values(ascending=False)
                 .head(10))
        for ticker, val in top10.items():
            ws.cell(row=row, column=1, value=ticker)
            val_cell = ws.cell(row=row, column=2, value=float(val))
            val_cell.number_format = "$#,##0"
            row += 1


def _build_monthly_summary(ws, df: pd.DataFrame):
    ws.title = "Monthly Summary"
    if df.empty:
        ws.cell(row=1, column=1, value="No data available")
        return

    df2 = df.copy()
    df2["month"] = df2["filed_date"].dt.to_period("M").astype(str)
    monthly = (df2.groupby("month")
               .agg(buy_count=("total_value", "count"),
                    total_value=("total_value", "sum"),
                    avg_value=("total_value", "mean"),
                    unique_companies=("ticker", "nunique"))
               .reset_index()
               .sort_values("month", ascending=False))

    headers = ["Month", "# Buys", "Total Value (USD)", "Avg Value (USD)", "Unique Companies"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, record in enumerate(monthly.itertuples(index=False), 2):
        ws.cell(row=row_idx, column=1, value=record.month)
        ws.cell(row=row_idx, column=2, value=int(record.buy_count))
        v = ws.cell(row=row_idx, column=3, value=float(record.total_value))
        v.number_format = "$#,##0"
        a = ws.cell(row=row_idx, column=4, value=float(record.avg_value))
        a.number_format = "$#,##0"
        ws.cell(row=row_idx, column=5, value=int(record.unique_companies))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def export_to_excel(df: pd.DataFrame, scanned_count: int, output_path: str = OUTPUT_FILE):
    logger.info("Building Excel report: %s", output_path)
    # Ensure required columns exist even when df is empty
    if df.empty:
        df = pd.DataFrame(columns=DISPLAY_COLUMNS)
    else:
        for col in DISPLAY_COLUMNS:
            if col not in df.columns:
                df[col] = None
        df["cluster_buy"] = df["cluster_buy"].fillna(False).astype(bool)
        df["source"] = df["source"].fillna("Form 4")
    wb = Workbook()

    # Sheet 1: Dashboard
    ws_dash = wb.active
    _build_dashboard(ws_dash, df, scanned_count)

    # Sheet 2: All Buys
    ws_all = wb.create_sheet("All Buys")
    ws_all.title = "All Buys"
    _write_df_to_sheet(ws_all, df)

    # Sheet 3: Cluster Buys
    ws_cluster = wb.create_sheet("Cluster Buys")
    ws_cluster.title = "Cluster Buys"
    if df.empty or "cluster_buy" not in df.columns:
        cluster_df = pd.DataFrame(columns=DISPLAY_COLUMNS)
    else:
        cluster_df = df[df["cluster_buy"].astype(bool)].copy()
        if not cluster_df.empty:
            cluster_df = cluster_df.sort_values(["ticker", "transaction_date"])
        cluster_df = cluster_df.reset_index(drop=True)
    _write_df_to_sheet(ws_cluster, cluster_df)

    # Sheet 4: Monthly Summary
    ws_monthly = wb.create_sheet("Monthly Summary")
    _build_monthly_summary(ws_monthly, df)

    wb.save(output_path)
    logger.info("Excel saved to %s", output_path)
    return output_path
